import win32com.client as win32

from openpyxl import load_workbook

import os


def none_to_zero(val):
    return (val, 0)[val is None]


def get_coordinate(row: int, col: str):
    return col + str(row)


class ReportCreator:
    LOOKUP_TABLE = [
        ('J3', 'D'),

        ('D4', 'C'),
        
        ('V21', 'I'),
        ('V22', 'J'),
        ('V23', 'E'),
        
        ('V28', 'K'),

        ('V31', 'N'),
        ('V32', 'O'),
        ('V33', 'P'),

        ('V36', 'L'),
        ('V37', 'M'),

        ('V51', 'R'),
        ('V52', 'S'),
        ('V53', 'V'),

        ('V56', 'T'),
        ('V57', 'U'),

        ('V79', 'Z'),
        ('V80', 'AA'),
        ('V81', 'X'),
    ]

    def __init__(self, path_to_source, path_to_template, path_to_out_folder):       
        self.excel = win32.gencache.EnsureDispatch('Excel.Application')
        
        self.old_source = (path_to_source[-4] == '.')
        if self.old_source:
            path_to_source = self.xls2xlsx(path_to_source)
            # raise Exception('File with the xlsx extension was expected.')
        self.path_to_source = path_to_source

        self.old_template = (path_to_template[-4] == '.')
        if self.old_template:
            path_to_template = self.xls2xlsx(path_to_template)
            # raise Exception('File with the xlsx extension was expected.')
        self.path_to_template = path_to_template

        self.path_to_out_folder = path_to_out_folder

        # source workbook
        self.source_wb = load_workbook(self.path_to_source)
        self.source_ws = self.source_wb.active

        self.total_rows = 0
        for i in range(1, self.source_ws.max_row):
            if self.source_ws.cell(row=i, column=1).value == 'ВСЕГО:':
                self.total_rows = i
                break

        # get last word from the first line
        self.current_year = self.source_ws['A1'].value.split()[-1]

    def xls2xlsx(self, abspath_to_xls, new_name=None):
        abspath_to_xlsx = (new_name, abspath_to_xls + 'x')[new_name is None]
        if not os.path.isfile(abspath_to_xlsx):
            wb = self.excel.Workbooks.Open(abspath_to_xls)
            wb.SaveAs(abspath_to_xlsx, FileFormat=51)  # FileFormat=51 is for .xlsx extension
            wb.Close()
        return abspath_to_xlsx

    def xlsx2xls(self, abspath_to_xlsx, new_name=None):
        abspath_to_xls = (new_name, abspath_to_xlsx[:-1])[new_name is None]
        if not os.path.isfile(abspath_to_xls):
            wb = self.excel.Workbooks.Open(abspath_to_xlsx)
            wb.SaveAs(abspath_to_xls, FileFormat=56)  # FileFormat=56 is for .xls extension
            wb.Close()
        return abspath_to_xls

    def create_file_by_template(self, file_path):
        if not os.path.isfile(file_path):
            wb = self.excel.Workbooks.Open(self.path_to_template)
            wb.SaveAs(file_path, FileFormat=51)  # FileFormat=51 is for .xlsx extension
            wb.Close()

    def apply_replacement(self, result_ws, result_coord, source_coord, rule=lambda x: x):
        val = none_to_zero(self.source_ws[source_coord].value)
        result_ws[result_coord].value = rule(val)

    def create_report_file(self, row):
        name = self.source_ws[get_coordinate(row, "B")].value
        name = name.replace('"', ' ')
        name = ' '.join(name.split())

        report_name = f'{self.source_ws[get_coordinate(row, "A")].value}_' \
                      f'{name}_' \
                      f'{self.current_year}.xlsx'

        report_path = os.path.join(self.path_to_out_folder, report_name)

        self.create_file_by_template(report_path)

        result_wb = load_workbook(report_path)
        result_ws = result_wb.active

        for replacement in self.LOOKUP_TABLE:
            result_coord = replacement[0]
            source_coord = replacement[1]

            if len(replacement) > 2:
                rule = replacement[2]
                self.apply_replacement(result_ws, result_coord, get_coordinate(row, source_coord), rule)
            else:
                self.apply_replacement(result_ws, result_coord, get_coordinate(row, source_coord))

        result_wb.save(report_path)
        result_wb.close()

        self.xlsx2xls(report_path)

        os.remove(report_path)

    def __del__(self):
        if self.old_source:
            os.remove(self.path_to_source)

        if self.old_template:
            os.remove(self.path_to_template)

        self.excel.Application.Quit()
